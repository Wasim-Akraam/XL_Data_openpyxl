import openpyxl
path = "../Data/Test.xlsx"

class Xlutility:

    def max_row(self,filename, sheetname):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        return sheet.max_row

    def max_col(self,filename,sheetname):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        return sheet.max_column


    def readdata(self,filename,sheetname,row,col):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        return sheet.cell(row=row,column=col).value
    def writedata(self,filename,sheetname,row,col,data):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        sheet.cell(row=row, column=col).value=data
        workbook.save(filename)


d='successfull'

obj=Xlutility()
obj.writedata(path,'Sheet1',row=1,col=4,data=d)
print(obj.readdata(path,'Sheet1',row=1,col=4))




